"""
Daily Update Excel Automation
===============================
Automatically generates or appends to a daily Excel report.
Columns: Date | Task/Activity | Status | Assigned To | Comments/Notes
"""

import os
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# CONFIGURATION — edit these values as needed
# ─────────────────────────────────────────────
OUTPUT_DIR   = os.path.dirname(os.path.abspath(__file__))   # same folder as script
REPORT_NAME  = "daily_updates.xlsx"                          # output filename

# ---------------------------------------------------------------------------
# Daily tasks to add  ← UPDATE THIS LIST each day (or connect to a DB/API)
# ---------------------------------------------------------------------------
TODAYS_TASKS = [
    {
        "task":      "Review project requirements",
        "status":    "Done",
        "assigned":  "Alice",
        "comments":  "All requirements confirmed with client",
    },
    {
        "task":      "Fix login page bug",
        "status":    "In Progress",
        "assigned":  "Bob",
        "comments":  "JWT token expiry issue identified",
    },
    {
        "task":      "Deploy to staging server",
        "status":    "Pending",
        "assigned":  "Charlie",
        "comments":  "Waiting for QA sign-off",
    },
]
# ---------------------------------------------------------------------------

# ─────────────────────────────────────────────
# STYLE CONSTANTS
# ─────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")      # dark blue
ALT_ROW_FILL  = PatternFill("solid", fgColor="D6E4F0")      # light blue
DONE_FILL     = PatternFill("solid", fgColor="C6EFCE")      # green
PROGRESS_FILL = PatternFill("solid", fgColor="FFEB9C")      # yellow
PENDING_FILL  = PatternFill("solid", fgColor="FFC7CE")      # red/pink

HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT     = Font(name="Calibri", size=10)
CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT          = Alignment(horizontal="left",   vertical="center", wrap_text=True)

THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = [
    ("Date",          18),
    ("Task / Activity", 40),
    ("Status",        18),
    ("Assigned To",   20),
    ("Comments / Notes", 45),
]

STATUS_FILLS = {
    "done":        DONE_FILL,
    "in progress": PROGRESS_FILL,
    "pending":     PENDING_FILL,
}


# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def _apply_header(ws):
    """Write and style the header row."""
    for col_idx, (col_name, col_width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER
        cell.border    = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    ws.row_dimensions[1].height = 22


def _style_data_row(ws, row_num, status_str):
    """Apply background and style to a data row."""
    status_key  = status_str.strip().lower()
    status_fill = STATUS_FILLS.get(status_key, PatternFill())   # default = no fill
    alt_fill    = ALT_ROW_FILL if row_num % 2 == 0 else PatternFill()

    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.font   = BODY_FONT
        cell.border = BORDER
        ws.row_dimensions[row_num].height = 18

        # Status column gets colour-coded fill; others get alternating rows
        if col_idx == 3:
            cell.fill      = status_fill
            cell.alignment = CENTER
        elif col_idx == 1:
            cell.alignment = CENTER
            cell.fill      = alt_fill
        else:
            cell.alignment = LEFT
            cell.fill      = alt_fill


def _freeze_and_filter(ws):
    """Freeze header row and enable auto-filter."""
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def generate_report(tasks: list[dict]):
    filepath = os.path.join(OUTPUT_DIR, REPORT_NAME)
    today    = date.today().strftime("%Y-%m-%d")

    # Load existing workbook or create a new one
    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        ws = wb.active
        next_row = ws.max_row + 1
        print(f"[INFO] Appending to existing file: {filepath}")
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Daily Updates"
        _apply_header(ws)
        next_row = 2
        print(f"[INFO] Creating new file: {filepath}")

    # Write today's tasks
    for task in tasks:
        ws.cell(row=next_row, column=1, value=today)
        ws.cell(row=next_row, column=2, value=task.get("task",     ""))
        ws.cell(row=next_row, column=3, value=task.get("status",   "Pending"))
        ws.cell(row=next_row, column=4, value=task.get("assigned", ""))
        ws.cell(row=next_row, column=5, value=task.get("comments", ""))
        _style_data_row(ws, next_row, task.get("status", "Pending"))
        next_row += 1

    _freeze_and_filter(ws)

    wb.save(filepath)
    print(f"[SUCCESS] Report saved → {filepath}  ({len(tasks)} task(s) added for {today})")


if __name__ == "__main__":
    generate_report(TODAYS_TASKS)
